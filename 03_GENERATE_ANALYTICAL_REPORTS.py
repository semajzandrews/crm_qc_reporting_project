import openpyxl
import os
import re
import json
import tkinter as tk
from tkinter import simpledialog, messagebox
from pdfminer.high_level import extract_text
from datetime import datetime

# Relative path configuration
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TARGETS_FILE = os.path.join(BASE_DIR, "targets.json")

# Load Configuration from Script 01
if not os.path.exists(TARGETS_FILE):
    print(f"Error: {os.path.basename(TARGETS_FILE)} not found. Please run Script 01 first.")
    exit(1)

with open(TARGETS_FILE, "r") as f:
    config_meta = json.load(f)

HS_DIR = config_meta.get("hs_dir")
SF_DIR = config_meta.get("sf_dir")
TEMPLATE_PATH = config_meta.get("template_path")
TARGETS_DICT = config_meta.get("matches", {})

# Results are always localized to the Template directory
RESULTS_DIR = os.path.join(os.path.dirname(TEMPLATE_PATH), "QA_ANALYTICS_RESULTS")
OUTPUT_EXCEL = os.path.join(RESULTS_DIR, 'QA_ANALYTICS_REPORT_FINAL.xlsx')
LOG_FILE = os.path.join(RESULTS_DIR, "QA_TECHNICAL_EVIDENCE.md")

SHEET_NAME = 'QA Report Test Tracker'
TESTER_NAME = "Semaj Andrews"

if not os.path.exists(RESULTS_DIR):
    os.makedirs(RESULTS_DIR)

# Identification Anchors - REFINED FOR OVERFLOW AWARENESS
SECTIONS = [
    {'key': 'Summary Page', 'marker': 'Year Over Year Comparison of Calls', 'next_marker': 'Calls by Site'},
    {'key': 'Site Page', 'marker': 'Calls by Site', 'next_marker': 'Calls by Day of Week'},
    {'key': 'Day of Week', 'marker': 'Calls by Day of Week', 'next_marker': 'Calls by Hour of Day'},
    {'key': 'Hour of Day', 'marker': 'Calls by Hour of Day', 'next_marker': 'Calls by Outcome'},
    {'key': 'Outcome', 'marker': 'Calls by Outcome', 'next_marker': 'Calls by Diagnosis'},
    {'key': 'Diagnosis', 'marker': 'Calls by Diagnosis', 'next_marker': None}
]

def clean_text(text, marker_to_purge=None):
    """Normalize whitespace and remove redundant headers to handle multi-page wraps."""
    if not text: return ""
    
    # 1. Strip redundant headers that repeat on page wraps
    if marker_to_purge:
        text = re.sub(re.escape(marker_to_purge), '', text, flags=re.IGNORECASE)
    
    # 2. Collapse whitespace
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def get_section_text(text, config):
    """
    Slices PDF text based on greedy anchors to capture multi-page tables.
    Uses 'Reach Next Expected Text' logic.
    """
    start_marker = config['marker']
    end_marker = config['next_marker']
    
    try:
        # Find the very first occurrence of the start marker
        start_match = re.search(re.escape(start_marker), text, re.IGNORECASE)
        if not start_match:
            return None
        
        start_idx = start_match.start()
        
        # Search for the next expected marker ONLY after the start index
        if end_marker:
            end_match = re.search(re.escape(end_marker), text[start_idx + len(start_marker):], re.IGNORECASE)
            if end_match:
                end_idx = start_idx + len(start_marker) + end_match.start()
            else:
                # If next marker not found, we look for any subsequent marker in the list
                # to prevent capturing the entire rest of the document
                end_idx = len(text)
                for sec in SECTIONS:
                    if sec['marker'] == start_marker: continue
                    alt_match = re.search(re.escape(sec['marker']), text[start_idx + len(start_marker):], re.IGNORECASE)
                    if alt_match:
                        found_idx = start_idx + len(start_marker) + alt_match.start()
                        if found_idx < end_idx: end_idx = found_idx
        else:
            end_idx = len(text)
            
        return text[start_idx:end_idx].strip()
    except Exception:
        return None

def process_client_analysis(sheet, row_idx, col_map, client_name, hs_path, sf_path, client_lines):
    print(f"--- PROCESSING: {client_name} ---")
    client_lines.append(f"\n### Client: {client_name}")
    
    if not os.path.exists(hs_path) or not os.path.exists(sf_path):
        msg = "Error: File paths could not be verified."
        print(f"   [{msg}]")
        client_lines.append(f"- {msg}")
        return False

    # 1. Binary comparison check
    with open(hs_path, 'rb') as f1, open(sf_path, 'rb') as f2:
        if f1.read() == f2.read():
            print("   Status: Exact binary match identified.")
            client_lines.append("- **Overall Result: 0** (Verified Binary Match)")
            sheet.cell(row=row_idx, column=col_map.get('Tester', 3)).value = TESTER_NAME
            report_col = col_map.get('Report ') or col_map.get('Report', 4)
            sheet.cell(row=row_idx, column=report_col).value = client_name
            for sec in SECTIONS:
                col_idx = col_map.get(sec['key'])
                if col_idx: sheet.cell(row=row_idx, column=col_idx).value = 0
            res_col = col_map.get('Test Result')
            if res_col: sheet.cell(row=row_idx, column=res_col).value = 0
            return True

    # 2. Sectional content analysis
    try:
        text_hs = extract_text(hs_path)
        text_sf = extract_text(sf_path)
    except Exception as e:
        msg = f"Error: Data extraction failed - {e}"
        print(f"   [{msg}]")
        client_lines.append(f"- {msg}")
        return False

    any_failure = False
    summary_present_in_both = False # Track if Summary Page exists to apply cascading logic later
    sheet.cell(row=row_idx, column=col_map.get('Tester', 3)).value = TESTER_NAME
    report_col = col_map.get('Report ') or col_map.get('Report', 4)
    sheet.cell(row=row_idx, column=report_col).value = client_name

    for section in SECTIONS:
        hs_raw = get_section_text(text_hs, section)
        sf_raw = get_section_text(text_sf, section)
        
        # LOGIC: If a section is missing from one but present in the other
        if hs_raw is None and sf_raw is None:
            if section['key'] == 'Summary Page':
                # Summary Page: Both missing is Acceptable (Permanent 0)
                result = 0
                reason = "Section missing in both sources (Acceptable)"
                summary_present_in_both = False
            else:
                # All other sections MUST be present. Missing = detection error = flag it.
                result = 1
                reason = "Section missing in both sources (Detection Error)"
        elif hs_raw is None or sf_raw is None:
            result = 1
            reason = "Section presence mismatch"
            if section['key'] == 'Summary Page': summary_present_in_both = False # Already failed
        else:
            if section['key'] == 'Summary Page': summary_present_in_both = True
            
            # STRIP REPEATED HEADERS BEFORE COMPARING
            clean_hs = clean_text(hs_raw, section['marker'])
            clean_sf = clean_text(sf_raw, section['marker'])
            
            # VOLUME GUARD: If data was found but it is just empty whitespace or single words
            # Adjusted threshold to 10 chars to be safe but catch empty tables
            if len(clean_hs) < 10 and len(clean_sf) < 10:
                result = 0
                reason = "Empty table match"
            elif len(clean_hs) < 10 or len(clean_sf) < 10:
                result = 1
                reason = "Content volume mismatch (One side empty)"
            else:
                result = 0 if clean_hs == clean_sf else 1
                reason = "Data Match" if result == 0 else "Data Discrepancy Identified"
            
        if result == 1: any_failure = True
        client_lines.append(f"- **{section['key']}**: {result} ({reason})")
        
        col_idx = col_map.get(section['key'])
        if col_idx: sheet.cell(row=row_idx, column=col_idx).value = result
        
    # 3. Final Integrity Guard: Summary Page inherits sub-page failures
    # This overrides the result ONLY if the Summary Page was actually present in both.
    # If it was missing in both (Acceptable), it stays 0 regardless of other failures.
    summary_key = 'Summary Page'
    summary_col = col_map.get(summary_key)
    if any_failure and summary_col and summary_present_in_both:
        sheet.cell(row=row_idx, column=summary_col).value = 1
        found_summary_line = False
        for i, line in enumerate(client_lines):
            if f"**{summary_key}**: 0" in line:
                client_lines[i] = f"- **{summary_key}**: 1 (Inferred mismatch: Supplemental data errors detected)"
                found_summary_line = True
                break
        if not found_summary_line:
             # If summary was already 1, we don't need to change the log, but the logic holds.
             pass
    
    test_result_col = col_map.get('Test Result')
    overall = 1 if any_failure else 0
    if test_result_col: sheet.cell(row=row_idx, column=test_result_col).value = overall
    client_lines.append(f"- **Analytical Verdict**: {overall}")
    return True

def generate_final_analytics():
    """Generate Excel report and text log with batching and physical verification."""
    root = tk.Tk()
    root.withdraw()
    
    if not os.path.exists(OUTPUT_EXCEL):
        for name in TARGETS_DICT: TARGETS_DICT[name]['status_excel'] = 'pending'
    else:
        wb_check = openpyxl.load_workbook(OUTPUT_EXCEL)
        sheet_check = wb_check[SHEET_NAME]
        header_row = 3
        report_col_idx = 4
        for idx, cell in enumerate(sheet_check[header_row]):
            if cell.value and "Report" in str(cell.value):
                report_col_idx = idx + 1
                break
        existing_agencies = {str(sheet_check.cell(row=r, column=report_col_idx).value).strip() 
                             for r in range(header_row + 1, sheet_check.max_row + 1) 
                             if sheet_check.cell(row=r, column=report_col_idx).value}
        for name in TARGETS_DICT:
            TARGETS_DICT[name]['status_excel'] = 'completed' if name in existing_agencies else 'pending'

    ready_targets = {k: v for k, v in TARGETS_DICT.items() 
                     if v.get('status_pdf') == 'completed' and v.get('status_excel', 'pending') == 'pending'}
    total_ready = len(ready_targets)
    
    if total_ready == 0:
        if any(v.get('status_pdf', 'pending') == 'pending' for v in TARGETS_DICT.values()):
            messagebox.showwarning("Prerequisite Not Met", "No files ready. Run Script 02 first.")
        else:
            messagebox.showinfo("Complete", "All processed!")
        return

    batch_size = simpledialog.askinteger("Batch Size", f"Files Ready: {total_ready}\nHow many?", 
                                       initialvalue=total_ready, minvalue=1, maxvalue=total_ready)
    if not batch_size: return

    target_excel = OUTPUT_EXCEL if os.path.exists(OUTPUT_EXCEL) else TEMPLATE_PATH
    wb = openpyxl.load_workbook(target_excel)
    sheet = wb[SHEET_NAME]
    header_row = 3
    col_map = {str(cell.value).strip(): idx + 1 for idx, cell in enumerate(sheet[header_row]) if cell.value}

    report_col_idx = col_map.get('Report ') or col_map.get('Report', 4)
    current_row = header_row + 1
    while sheet.cell(row=current_row, column=report_col_idx).value: current_row += 1

    processed_count = 0
    for client_name, paths in ready_targets.items():
        if processed_count >= batch_size: break
        client_lines = []
        if process_client_analysis(sheet, current_row, col_map, client_name, paths['hs'], paths['sf'], client_lines):
            TARGETS_DICT[client_name]['status_excel'] = 'completed'
            config_meta['matches'] = TARGETS_DICT
            with open(TARGETS_FILE, "w") as f: json.dump(config_meta, f, indent=4)
            wb.save(OUTPUT_EXCEL)
            with open(LOG_FILE, "a") as f: f.write("\n".join(client_lines) + "\n")
            print(f"   [FINALIZED] {client_name} (Row {current_row})")
            current_row += 1
            processed_count += 1

    messagebox.showinfo("Batch Complete", f"Processed: {processed_count}")

if __name__ == "__main__":
    generate_final_analytics()
